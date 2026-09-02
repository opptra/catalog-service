"""Flatfile template helpers — paths, parsing, and field validation.

Reusable from the job service; no FastAPI or repository imports.
"""

import csv
import io
import re
from pathlib import PurePosixPath
from typing import Any
from uuid import UUID

from openpyxl import load_workbook

from core.exceptions import FlatfileValidationError

SIGNED_URL_TTL_SECONDS = 3600

_SAFE_TEMPLATE_FILENAME = re.compile(r"^[\w.\- ()]+\.(csv|xlsx|xls)$", re.IGNORECASE)
_SAFE_IMAGE_FILENAME = re.compile(
    r"^[\w.\- ()]+\.(jpe?g|png|gif|webp|bmp|tiff?)$",
    re.IGNORECASE,
)


def template_object_key(job_external_id: UUID, filename: str) -> str:
    return f"jobs/{job_external_id}/input/template/{filename}"


def listing_template_object_key(marketplace_external_id: UUID, category_external_id: UUID) -> str:
    """Stable GCS key for the Amazon listing template for a given category × marketplace.

    Path follows the same entity/id/asset convention as product images:
    ``marketplaces/{marketplace_id}/categories/{category_id}/listing-template/template.xlsx``
    """
    return (
        f"listing-templates"
        f"/marketplaces/{marketplace_external_id}"
        f"/categories/{category_external_id}"
        f"/template.xlsx"
    )


def manifest_object_key(job_external_id: UUID) -> str:
    return f"jobs/{job_external_id}/input/manifest.json"


def product_image_prefix(sku_id: str) -> str:
    return f"products/{sku_id}/assets/images/"


def product_image_object_key(sku_id: str, filename: str) -> str:
    return f"{product_image_prefix(sku_id)}{filename}"


def safe_sku_id(sku_id: str) -> str:
    """Business SKU string (folder name), not sku_master.id."""
    text = sku_id.strip()
    if not text:
        raise FlatfileValidationError("SKU is required")
    if "/" in text or "\\" in text or text in {".", ".."}:
        raise FlatfileValidationError(f"Invalid SKU: {sku_id!r}")
    name = PurePosixPath(text).name
    if name != text:
        raise FlatfileValidationError(f"Invalid SKU: {sku_id!r}")
    return name


def safe_template_filename(filename: str) -> str:
    name = PurePosixPath(filename).name
    if not _SAFE_TEMPLATE_FILENAME.match(name):
        raise FlatfileValidationError(f"Unsupported template filename: {filename!r}")
    return name


def safe_image_filename(filename: str) -> str:
    name = PurePosixPath(filename).name
    if not _SAFE_IMAGE_FILENAME.match(name):
        raise FlatfileValidationError(f"Unsupported image filename: {filename!r}")
    return name


def cell_str(value: Any) -> str:
    """Coerce template cells to strings — never keep ints/floats in attributes."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _is_legend_row(values: list[str]) -> bool:
    """True when every non-empty cell is Mandatory/Optional (template styling row)."""
    non_empty = [value for value in values if value]
    if not non_empty:
        return False
    return all(value.lower() in {"mandatory", "optional"} for value in non_empty)


def row_get(row: dict[str, str], name: str) -> str:
    """Return the cell for an exact header key — keys are never rewritten."""
    return row.get(name, "")


def parse_template_rows(data: bytes, *, filename: str) -> tuple[list[str], list[dict[str, str]]]:
    """Parse template into headers + string-valued attribute rows (keys kept as written)."""
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames is None:
            raise FlatfileValidationError("Template CSV has no header row")
        headers = [str(name).strip() for name in reader.fieldnames if name is not None]
        rows: list[dict[str, str]] = []
        first_data = True
        for raw in reader:
            row = {
                str(key).strip(): cell_str(value)
                for key, value in raw.items()
                if key is not None and str(key).strip()
            }
            if not any(row.values()):
                continue
            if first_data and _is_legend_row(list(row.values())):
                first_data = False
                continue
            first_data = False
            rows.append(row)
        return headers, rows

    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration as exc:
            raise FlatfileValidationError("Template spreadsheet is empty") from exc
        headers = [cell_str(cell) for cell in header_row]
        if not any(headers):
            raise FlatfileValidationError("Template spreadsheet has no header row")
        rows = []
        first_data = True
        for values in iterator:
            row = {
                headers[index]: cell_str(values[index] if index < len(values) else None)
                for index in range(len(headers))
                if headers[index]
            }
            if not any(row.values()):
                continue
            if first_data and _is_legend_row(list(row.values())):
                first_data = False
                continue
            first_data = False
            rows.append(row)
        return headers, rows

    raise FlatfileValidationError(f"Unsupported template type: {filename}")


def validate_mandatory_fields(
    headers: list[str],
    rows: list[dict[str, str]],
    mandatory_names: list[str],
) -> None:
    """Require exact ``SKU`` + category mandatory columns, with a value on every row."""
    header_set = set(headers)
    required_columns = ["SKU", *mandatory_names]
    seen: set[str] = set()
    unique_missing: list[str] = []
    for name in required_columns:
        if name in seen:
            continue
        seen.add(name)
        if name not in header_set:
            unique_missing.append(name)
    if unique_missing:
        raise FlatfileValidationError(
            "Template missing required column(s): " + ", ".join(unique_missing)
        )

    if not rows:
        raise FlatfileValidationError("Template has no data rows")

    errors: list[str] = []
    for index, row in enumerate(rows, start=1):
        sku_raw = row_get(row, "SKU")
        sku_label = sku_raw or f"row {index}"
        if not sku_raw:
            errors.append(f"{sku_label}: missing SKU value")
            continue
        for name in mandatory_names:
            if name == "SKU":
                continue
            if not row_get(row, name):
                errors.append(f"{sku_label}: missing mandatory “{name}”")
    if errors:
        preview = "; ".join(errors[:20])
        suffix = "…" if len(errors) > 20 else ""
        raise FlatfileValidationError(preview + suffix)


def build_sku_attributes(
    row: dict[str, str],
    *,
    sku_id: str,
    existing_attributes: dict[str, Any] | None = None,
) -> dict[str, str]:
    """Build flat string key/value attributes from the template row only.

    Keys are kept exactly as in the spreadsheet — no case/space/punctuation rewrites.
    Images live in GCS under a predictable path — they are not stored in attributes.
    Existing non-string values (e.g. nested JSON) are dropped on merge.
    Callers must pass the result through the category allowed-list filter before persist.
    """
    attributes: dict[str, str] = {}
    if existing_attributes:
        for key, value in existing_attributes.items():
            if isinstance(value, str) and key:
                attributes[key] = value
    for key, value in row.items():
        if not key:
            continue
        attributes[key] = cell_str(value)
    # Ensure the validated SKU cell is present under the exact key "SKU".
    attributes["SKU"] = sku_id
    return attributes
