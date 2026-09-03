"""Build listing_template_column configs from a marketplace listing workbook.

Discovers dropdowns from Excel list-validation formulas (Amazon) and from
``DropDownValuesForColumn*`` / Index allowed-value sheets (Flipkart). No
category-specific column names hard-coded.

Does not write to any database.
"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Template cell refs like B7 / CJ7 — ignore sheet-qualified refs separately.
_CELL_REF_RE = re.compile(r"(?<![A-Z$!'])\b([A-Z]{1,3})(\d+)\b")
_VLOOKUP_CELL_RE = re.compile(r"VLOOKUP\(\s*([A-Z]{1,3})(\d+)", re.IGNORECASE)
# Cascading named-range construction:
#   ...&"parent_stem.value."&VLOOKUP(...)&".child_suffix"
_CASCADE_NAMED_RANGE_RE = re.compile(
    r'"([^"]+\.value\.)".*?VLOOKUP.*?&"\.([^"]+)"',
    re.IGNORECASE | re.DOTALL,
)
_DROPDOWN_VALUES_SHEET_RE = re.compile(r"^DropDownValuesForColumn(\d+)$")


@dataclass(frozen=True)
class WorkbookLayout:
    """Offsets into the blank marketplace workbook (same as listing_template.metadata).

    All values come from the CLI / caller — no marketplace-specific defaults.
    Optional sheet names may be omitted when that workbook has no such sheet.
    """

    sheet_name: str
    header_label_row: int
    machine_key_row: int
    data_start_row: int
    valid_values_sheet: str | None = None
    dropdown_lists_sheet: str | None = None
    data_definitions_sheet: str | None = None


def _sql_str(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def _json_sql(obj: dict[str, Any]) -> str:
    return _sql_str(json.dumps(obj, ensure_ascii=False, separators=(",", ":")))


def _requiredness(raw: str | None) -> str:
    if raw and str(raw).strip().casefold() == "required":
        return "ALWAYS"
    return "OPTIONAL"


def _cell_text(value: object | None) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return ILLEGAL_CHARACTERS_RE.sub("", str(value)).strip()


def _looks_like_url_column(label: str, type_hint: str) -> bool:
    """Skip dropdown-sheet attach on URL/image-link columns (Flipkart type row)."""
    hint = type_hint.casefold()
    if hint == "url":
        return True
    return "url" in label.casefold()


def _workbook_from_xls(path: Path) -> Workbook:
    """Load Excel 97-2003 .xls into an openpyxl workbook (values only)."""
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError(
            f"Cannot read {path.name}: Flipkart .xls workbooks need xlrd. "
            "Install it in the server venv (see server/requirements.txt)."
        ) from exc

    book = xlrd.open_workbook(str(path))
    wb = Workbook()
    default = wb.active
    for sheet_index, name in enumerate(book.sheet_names()):
        source = book.sheet_by_index(sheet_index)
        if sheet_index == 0:
            ws = default
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)
        for row in range(source.nrows):
            for col in range(source.ncols):
                text = _cell_text(source.cell_value(row, col))
                if text:
                    ws.cell(row + 1, col + 1, text)
    return wb


def _open_listing_workbook(path: Path) -> Workbook:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return load_workbook(path, read_only=False, data_only=False, keep_vba=True)
    if suffix == ".xls":
        return _workbook_from_xls(path)
    raise ValueError(
        f"Unsupported listing workbook type {path.suffix!r}. Expected .xlsx, .xlsm, or .xls"
    )


def _dropdown_values_by_column_sheets(wb: Workbook) -> dict[int, list[str]]:
    """Flipkart ``DropDownValuesForColumnN`` sheets; N is 0-based column index."""
    out: dict[int, list[str]] = {}
    for name in wb.sheetnames:
        match = _DROPDOWN_VALUES_SHEET_RE.match(name)
        if not match:
            continue
        excel_col = int(match.group(1)) + 1
        ws = wb[name]
        values: list[str] = []
        for row in range(1, (ws.max_row or 0) + 1):
            text = _cell_text(ws.cell(row, 1).value)
            if text:
                values.append(text)
        if values:
            out[excel_col] = values
    return out


def _index_sheet_allowed_values(wb: Workbook) -> dict[str, list[str]]:
    """Flipkart Index sheet: row 2 = field names, row 3+ = allowed values.

    Duplicate header names (parent vs listing copies) keep the first list.
    """
    if "Index" not in wb.sheetnames:
        return {}
    ws = wb["Index"]
    max_col = ws.max_column or 0
    row1 = [_cell_text(ws.cell(1, col).value).casefold() for col in range(1, max_col + 1)]
    if "allowed values" not in row1:
        return {}
    out: dict[str, list[str]] = {}
    for col in range(1, max_col + 1):
        name = _cell_text(ws.cell(2, col).value)
        if not name or name in out:
            continue
        values: list[str] = []
        for row in range(3, (ws.max_row or 0) + 1):
            text = _cell_text(ws.cell(row, col).value)
            if text:
                values.append(text)
        if values:
            out[name] = values
    return out


def _parse_static_list(formula: str) -> list[str] | None:
    f = formula.strip()
    if not f:
        return None
    if f.startswith('"') and f.endswith('"'):
        inner = f[1:-1]
        if "," in inner:
            return [p.strip() for p in inner.split(",") if p.strip()]
        if inner:
            return [inner]
    return None


def _strip_sheet_refs(formula: str) -> str:
    """Remove 'Sheet'!A1 and Sheet!A1 so cell-ref scans stay on the Template row."""
    cleaned = re.sub(r"'[^']+'![^,&\)]+", "", formula)
    return re.sub(r"[A-Za-z_][\w.]*![^,&\)]+", "", cleaned)


def _load_valid_values_by_label(
    wb: Workbook,
    *,
    sheet_name: str | None,
) -> dict[str, list[str]]:
    """Map Template local label → allowed values (Valid Values sheet)."""
    if not sheet_name or sheet_name not in wb.sheetnames:
        return {}
    vv = wb[sheet_name]
    out: dict[str, list[str]] = {}
    for row in range(1, (vv.max_row or 0) + 1):
        header = vv.cell(row, 2).value
        if not header or " - [" not in str(header):
            continue
        label = str(header).split(" - [", 1)[0].strip()
        values: list[str] = []
        col = 3
        while True:
            cell = vv.cell(row, col).value
            if cell is None:
                break
            text = str(cell).strip()
            if text:
                values.append(text)
            col += 1
        if values:
            out[label] = values
    return out


def _dropdown_display_to_token(
    wb: Workbook,
    *,
    sheet_name: str | None,
) -> dict[str, str]:
    if not sheet_name or sheet_name not in wb.sheetnames:
        return {}
    dl = wb[sheet_name]
    mapping: dict[str, str] = {}
    for row in range(1, (dl.max_row or 0) + 1):
        display = dl.cell(row, 1).value
        token = dl.cell(row, 2).value
        if display and token:
            mapping[str(display).strip()] = str(token).strip()
    return mapping


def _resolve_defined_name(wb: Workbook, name: str) -> list[str]:
    if name not in wb.defined_names:
        return []
    dests = list(wb.defined_names[name].destinations)
    if not dests:
        return []
    sheet_name, coord = dests[0]
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    values: list[str] = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col).value
            if cell is None:
                continue
            text = str(cell).strip()
            if text:
                values.append(text)
    return values


def _product_type_values(values_by_label: dict[str, list[str]]) -> list[str]:
    return list(values_by_label.get("Product Type") or [])


def _product_type_token(values_by_label: dict[str, list[str]]) -> str:
    pts = _product_type_values(values_by_label)
    if pts:
        return pts[0].replace("-", "_").replace(" ", "")
    return "PRODUCT"


def _data_definition_requiredness(
    wb: Workbook,
    *,
    sheet_name: str | None,
) -> dict[str, str]:
    if not sheet_name or sheet_name not in wb.sheetnames:
        return {}
    dd = wb[sheet_name]
    out: dict[str, str] = {}
    for row in range(4, (dd.max_row or 0) + 1):
        field = dd.cell(row, 2).value
        if not field:
            continue
        out[str(field).strip()] = _requiredness(dd.cell(row, 6).value)
    return out


def _list_validations_by_column(ws: Worksheet) -> dict[int, str]:
    by_col: dict[int, str] = {}
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue
        formula = str(dv.formula1 or "")
        for ref in str(dv.sqref).split():
            match = re.match(r"([A-Z]+)", ref)
            if not match:
                continue
            col = column_index_from_string(match.group(1))
            by_col[col] = formula
    return by_col


def _immediate_parent_column(
    *,
    formula: str,
    column_index: int,
    data_row: int,
) -> int | None:
    """Parent Template column this list is gated on (from formula cell refs).

    Priority:
      1) VLOOKUP(<col><row>, ...) → that column (cascading enum parent)
      2) Otherwise any same-row Template cell ref (typically Product Type)
    """
    vlookup = _VLOOKUP_CELL_RE.search(formula)
    if vlookup and int(vlookup.group(2)) == data_row:
        parent = column_index_from_string(vlookup.group(1))
        if parent != column_index:
            return parent

    cleaned = _strip_sheet_refs(formula)
    parents: list[int] = []
    for match in _CELL_REF_RE.finditer(cleaned):
        col = column_index_from_string(match.group(1))
        row = int(match.group(2))
        if row != data_row or col == column_index:
            continue
        if col not in parents:
            parents.append(col)
    if not parents:
        return None
    return parents[0]


def _cascaded_values_by_parent(
    wb: Workbook,
    *,
    formula: str,
    parent_values: list[str],
    product_tokens: list[str],
    dropdown_lists_sheet: str | None,
) -> dict[str, list[str]] | None:
    """Build parent_value → child options from named ranges encoded in the formula."""
    match = _CASCADE_NAMED_RANGE_RE.search(formula)
    if not match:
        return None
    mid_prefix = match.group(1)  # e.g. league_name.value.
    child_suffix = match.group(2)  # e.g. team_namemarketplace_id...
    display_to_token = _dropdown_display_to_token(wb, sheet_name=dropdown_lists_sheet)
    out: dict[str, list[str]] = {}
    for parent_val in parent_values:
        token = display_to_token.get(parent_val)
        if not token:
            continue
        values: list[str] = []
        for pt_token in product_tokens:
            name = f"{pt_token}{mid_prefix}{token}.{child_suffix}"
            values = _resolve_defined_name(wb, name)
            if values:
                break
        out[parent_val] = values
    return out


def build_columns(
    xlsm_path: Path,
    *,
    layout: WorkbookLayout,
    include_requiredness: bool = True,
) -> list[dict[str, Any]]:
    """Parse workbook → column dicts with resolve_stage + config.

    ``layout`` must be supplied by the caller (CLI flags / metadata) — no
    marketplace defaults are applied here.
    When ``include_requiredness`` is False, configs omit Data Definitions
    required/optional (listing-mapping uses category attribute_spec instead).
    """
    wb = _open_listing_workbook(xlsm_path)
    if layout.sheet_name not in wb.sheetnames:
        raise ValueError(f"Workbook needs sheet {layout.sheet_name!r}")

    ws = wb[layout.sheet_name]
    req_by_key: dict[str, str] = {}
    if include_requiredness:
        req_by_key = _data_definition_requiredness(wb, sheet_name=layout.data_definitions_sheet)
    values_by_label = _load_valid_values_by_label(wb, sheet_name=layout.valid_values_sheet)
    list_formulas = _list_validations_by_column(ws)
    dropdown_by_col = _dropdown_values_by_column_sheets(wb)
    dropdown_by_label = _index_sheet_allowed_values(wb)
    data_row = layout.data_start_row
    product_values = _product_type_values(values_by_label)
    product_tokens = (
        [v.replace("-", "_").replace(" ", "") for v in product_values]
        if product_values
        else [_product_type_token(values_by_label)]
    )

    max_col = ws.max_column or 0
    labels_by_index = {
        c: str(ws.cell(layout.header_label_row, c).value).strip()
        for c in range(1, max_col + 1)
        if ws.cell(layout.header_label_row, c).value
    }

    columns: list[dict[str, Any]] = []
    for col_index in range(1, max_col + 1):
        label = ws.cell(layout.header_label_row, col_index).value
        machine_key = ws.cell(layout.machine_key_row, col_index).value
        if label is None and machine_key is None:
            continue
        label_s = str(label).strip() if label else f"Column {col_index}"
        key_s = str(machine_key).strip() if machine_key else None

        formula = list_formulas.get(col_index)
        type_hint = key_s or ""
        sheet_enum: list[str] | None = None
        if formula is None:
            if not _looks_like_url_column(label_s, type_hint):
                sheet_enum = dropdown_by_col.get(col_index)
            if not sheet_enum:
                sheet_enum = dropdown_by_label.get(label_s)
        if formula is None and sheet_enum:
            config = {
                "fill_type": "ENUM",
                "label": label_s,
                "valid_values": sheet_enum,
            }
            if include_requiredness:
                config["requiredness"] = req_by_key.get(key_s or "", "OPTIONAL")
            columns.append(
                {
                    "column_index": col_index,
                    "depends_on": None,
                    "workbook_key": key_s,
                    "config": config,
                }
            )
            continue
        if formula is None:
            # Non-dropdown defaults to DIRECT_MAP. Product sets IMAGE / CONSTANT /
            # SKIP / AI_TEXT in DB (or a follow-up SQL) when known.
            config = {
                "fill_type": "DIRECT_MAP",
                "label": label_s,
            }
            if include_requiredness:
                config["requiredness"] = req_by_key.get(key_s or "", "OPTIONAL")
            columns.append(
                {
                    "column_index": col_index,
                    "depends_on": None,
                    # Parse-time only — used to match mapping CSV; not stored in config.
                    "workbook_key": key_s,
                    "config": config,
                }
            )
            continue

        parent_col = _immediate_parent_column(
            formula=formula,
            column_index=col_index,
            data_row=data_row,
        )
        # Parent pointer for fill is Excel column_index (stable for this template).
        depends_on: int | None = parent_col

        config = {
            "fill_type": "ENUM",
            "label": label_s,
        }
        if include_requiredness:
            config["requiredness"] = req_by_key.get(key_s or "", "OPTIONAL")
        if depends_on is not None:
            config["depends_on"] = depends_on

        parent_label = labels_by_index.get(parent_col or -1)
        parent_flat_values = (
            list(values_by_label.get(parent_label or "") or []) if parent_label else []
        )

        cascaded = None
        if depends_on is not None and _VLOOKUP_CELL_RE.search(formula):
            cascaded = _cascaded_values_by_parent(
                wb,
                formula=formula,
                parent_values=parent_flat_values,
                product_tokens=product_tokens,
                dropdown_lists_sheet=layout.dropdown_lists_sheet,
            )

        if cascaded is not None:
            config["valid_values_by_parent"] = cascaded
        else:
            values = values_by_label.get(label_s)
            if not values:
                values = _parse_static_list(formula)
            if not values:
                values = _resolve_defined_name(wb, formula.strip().strip('"'))
            if depends_on is not None and values and parent_flat_values:
                config["valid_values_by_parent"] = {
                    parent_val: list(values) for parent_val in parent_flat_values
                }
            elif depends_on is not None and values and product_values:
                config["valid_values_by_parent"] = {pt: list(values) for pt in product_values}
            elif values:
                config["valid_values"] = values
            else:
                config["valid_values"] = ["__UNRESOLVED_DROPDOWN__"]

        columns.append(
            {
                "column_index": col_index,
                "depends_on": depends_on,
                "workbook_key": key_s,
                "config": config,
            }
        )

    _assign_resolve_stages(columns)
    return columns


def _assign_resolve_stages(columns: list[dict[str, Any]]) -> None:
    index_to_col = {c["column_index"]: c for c in columns}
    depth_cache: dict[int, int] = {}

    def depth_of(col: dict[str, Any]) -> int:
        idx = col["column_index"]
        if idx in depth_cache:
            return depth_cache[idx]
        parent_index = col.get("depends_on")
        if parent_index is None:
            depth_cache[idx] = 0
            return 0
        parent = index_to_col.get(parent_index)
        if parent is None:
            depth_cache[idx] = 1
            return 1
        depth_cache[idx] = -1
        parent_depth = depth_of(parent)
        if parent_depth < 0:
            raise ValueError(f"Cycle in depends_on at column {idx}")
        depth = parent_depth + 1
        depth_cache[idx] = depth
        return depth

    for col in columns:
        col["resolve_stage"] = depth_of(col) + 1


def render_sql(columns: list[dict[str, Any]], *, xlsm_name: str) -> str:
    by_stage: dict[int, int] = defaultdict(int)
    by_fill: dict[str, int] = defaultdict(int)
    by_req: dict[str, int] = defaultdict(int)
    for col in columns:
        by_stage[col["resolve_stage"]] += 1
        by_fill[col["config"]["fill_type"]] += 1
        req = col["config"].get("requiredness")
        if req:
            by_req[req] += 1

    lines = [
        "-- Human-run only. Agent does not apply this against any database.",
        f"-- Generated from {xlsm_name} by ops.listing_mapping.generate_columns",
        "--",
        "-- Re-generate (from repo root, PYTHONPATH=ops:server):",
        "--   python -m listing_mapping.generate_columns \\",
        f"--     --xlsm /path/to/{xlsm_name} \\",
        "--     --marketplace <name> \\",
        "--     --sheet-name <sheet> \\",
        "--     --header-label-row <n> \\",
        "--     --machine-key-row <n> \\",
        "--     --data-start-row <n> \\",
        "--     --out tmp/sql/002_<category>_listing_columns.sql",
        "--",
        "-- Before apply: set :cm_id (category_marketplace.id) and ensure",
        "-- listing_template exists for that junction.",
        "--",
        f"-- Columns: {len(columns)} | fill_types={dict(by_fill)} |",
        f"-- requiredness={dict(by_req)} | resolve_stages={dict(sorted(by_stage.items()))}",
        "--",
        "-- resolve_stage = dependency depth starting at 1 (same stage concurrent; higher later).",
        "-- ENUM = dropdowns; DIRECT_MAP = everything else (map sources set later).",
        "-- depends_on / valid_values_by_parent discovered from Excel formulas.",
        "",
        "BEGIN;",
        "",
        "DELETE FROM listing_template_column",
        "WHERE listing_template_id = (",
        "    SELECT id FROM listing_template WHERE category_marketplace_id = :cm_id",
        ");",
        "",
        "INSERT INTO listing_template_column",
        "    (listing_template_id, column_index, resolve_stage, config)",
        "SELECT lt.id, v.column_index, v.resolve_stage, v.config::jsonb",
        "FROM listing_template lt",
        "CROSS JOIN (",
        "    VALUES",
    ]

    value_lines = [
        f"    ({col['column_index']}, {col['resolve_stage']}, {_json_sql(col['config'])})"
        for col in columns
    ]
    lines.append(",\n".join(value_lines))
    lines.extend(
        [
            ") AS v(column_index, resolve_stage, config)",
            "WHERE lt.category_marketplace_id = :cm_id;",
            "",
            "COMMIT;",
            "",
        ]
    )
    return "\n".join(lines)


def write_sql(
    xlsm_path: Path,
    out_path: Path,
    *,
    layout: WorkbookLayout,
) -> list[dict[str, Any]]:
    columns = build_columns(xlsm_path, layout=layout)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        render_sql(columns, xlsm_name=xlsm_path.name),
        encoding="utf-8",
    )
    return columns
