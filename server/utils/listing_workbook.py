"""openpyxl helpers for filling Amazon listing workbooks."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook

from dto.listing_config import ListingTemplateMetadata


def fill_workbook(
    template_bytes: bytes,
    *,
    metadata: ListingTemplateMetadata,
    rows: list[dict[int, str | None]],
) -> bytes:
    """Write SKU rows into a blank template and return workbook bytes.

    ``rows`` is a list of ``column_index → cell value`` maps (1-based Excel columns).
    Uses ``keep_vba=True`` so ``.xlsm`` macros survive when present.
    """
    buffer = io.BytesIO(template_bytes)
    try:
        workbook = load_workbook(buffer, keep_vba=True)
    except Exception:
        buffer.seek(0)
        workbook = load_workbook(buffer, keep_vba=False)

    if metadata.sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {metadata.sheet_name!r} not found in listing template")
    sheet = workbook[metadata.sheet_name]

    for row_offset, values in enumerate(rows):
        excel_row = metadata.data_start_row + row_offset
        for column_index, cell_value in values.items():
            if cell_value is None:
                continue
            sheet.cell(excel_row, column_index, cell_value)

    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def listing_output_object_key(job_external_id: Any, filename: str) -> str:
    """GCS key for a filled listing workbook under a generation job."""
    safe_name = filename.strip() or "listing.xlsx"
    return f"jobs/{job_external_id}/output/listing/{safe_name}"
