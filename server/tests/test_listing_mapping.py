"""Unit tests for ops/listing_mapping (mapping workbook + overlay + Amazon adapter)."""

from __future__ import annotations

from pathlib import Path

import pytest
from listing_mapping.mapping_workbook import (
    FillMode,
    ListingMapRow,
    MappingWorkbook,
    MarketplaceColumnRow,
    PimFieldRow,
    build_attribute_spec,
    parse_mapping_workbook,
)
from listing_mapping.marketplace import MarketplaceId, parse_marketplace_id
from listing_mapping.marketplace.registry import get_adapter
from listing_mapping.overlay import overlay_columns
from listing_mapping.render import render_mapping_sql
from openpyxl import Workbook

from utils.listing_template_columns import WorkbookLayout, build_columns

_REPO = Path(__file__).resolve().parents[2]
_MAP = _REPO / "ops" / "docs" / "listing_mapping_template.xlsx"


def _col(
    *,
    column_index: int,
    label: str,
    fill_type: str = "DIRECT_MAP",
    resolve_stage: int = 1,
    depends_on: int | None = None,
    valid_values: list[str] | None = None,
    valid_values_by_parent: dict[str, list[str]] | None = None,
) -> dict:
    config: dict = {"fill_type": fill_type, "label": label}
    if depends_on is not None:
        config["depends_on"] = depends_on
    if valid_values is not None:
        config["valid_values"] = valid_values
    if valid_values_by_parent is not None:
        config["valid_values_by_parent"] = valid_values_by_parent
    return {
        "column_index": column_index,
        "resolve_stage": resolve_stage,
        "depends_on": depends_on,
        "workbook_key": None,
        "config": config,
    }


def _mapping(
    listing_rows: list[ListingMapRow],
    *,
    pim_fields: list[PimFieldRow] | None = None,
) -> MappingWorkbook:
    return MappingWorkbook(
        pim_fields=pim_fields
        or [
            PimFieldRow("SKU", True),
            PimFieldRow("Color", True),
        ],
        marketplace_columns=[
            MarketplaceColumnRow(
                excel_row=2 + i,
                column_index=row.column_index,
                marketplace_column=f"col-{row.column_index}",
            )
            for i, row in enumerate(listing_rows)
        ],
        listing_rows=listing_rows,
    )


def _write_mapping(
    path: Path,
    *,
    rows: list[tuple[int, str, str, str, str, str]],
    pim: list[tuple[str, str]] | None = None,
    sheet: str = "amazon_mapping",
) -> None:
    wb = Workbook()
    ws_pim = wb.active
    ws_pim.title = "pim_contract"
    ws_pim.append(["pim_field", "requirement"])
    for field, req in pim or [("SKU", "Mandatory"), ("Color", "Optional")]:
        ws_pim.append([field, req])
    ws_map = wb.create_sheet(sheet)
    ws_map.append(
        [
            "column_index",
            "marketplace_column",
            "fill_mode",
            "pim_field",
            "generation",
            "constant_value",
            "status",
        ]
    )
    for row in rows:
        ws_map.append([*row, "OK"])
    wb.save(path)


def test_parse_marketplace_id() -> None:
    assert parse_marketplace_id("amazon") is MarketplaceId.AMAZON
    assert parse_marketplace_id("AMAZON") is MarketplaceId.AMAZON
    with pytest.raises(ValueError, match="Unknown marketplace"):
        parse_marketplace_id("SHOPIFY")


def test_amazon_adapter_default_layout() -> None:
    layout = get_adapter(MarketplaceId.AMAZON).workbook_layout()
    assert layout.sheet_name == "Template"
    assert layout.header_label_row == 4
    assert layout.machine_key_row == 5
    assert layout.data_start_row == 7
    assert layout.valid_values_sheet == "Valid Values"


def test_amazon_adapter_overrides() -> None:
    layout = get_adapter(MarketplaceId.AMAZON).workbook_layout(
        sheet_name="Custom",
        data_start_row=9,
    )
    assert layout.sheet_name == "Custom"
    assert layout.data_start_row == 9
    assert layout.header_label_row == 4


def test_flipkart_adapter_default_layout() -> None:
    layout = get_adapter(MarketplaceId.FLIPKART).workbook_layout()
    assert layout.sheet_name == "bedsheet"
    assert layout.header_label_row == 1
    assert layout.machine_key_row == 2
    assert layout.data_start_row == 5
    assert layout.valid_values_sheet is None


def test_flipkart_adapter_overrides() -> None:
    layout = get_adapter(MarketplaceId.FLIPKART).workbook_layout(sheet_name="curtain")
    assert layout.sheet_name == "curtain"
    assert layout.data_start_row == 5


def test_build_attribute_spec_injects_sku() -> None:
    spec = build_attribute_spec(
        [
            PimFieldRow("Color", True),
            PimFieldRow("Material", False),
        ]
    )
    assert spec["allowed"][0] == "SKU"
    assert "SKU" in spec["mandatory"]
    assert "Color" in spec["mandatory"]


def test_overlay_by_column_index_fill_modes() -> None:
    workbook = [
        _col(column_index=1, label="Seller SKU"),
        _col(
            column_index=2,
            label="Color",
            fill_type="ENUM",
            valid_values=["Black", "White"],
        ),
        _col(
            column_index=3,
            label="League",
            fill_type="ENUM",
            valid_values=["NFL"],
        ),
        _col(column_index=4, label="Item Name"),
        _col(column_index=5, label="Main Image"),
        _col(column_index=6, label="Notes"),
        _col(column_index=7, label="Unused"),
    ]
    mapping = _mapping(
        [
            ListingMapRow(1, FillMode.COPY_PIM, "SKU", None, None),
            ListingMapRow(2, FillMode.ENUM_FROM_PIM, "Color", None, None),
            ListingMapRow(3, FillMode.ENUM_AI, None, None, None),
            ListingMapRow(4, FillMode.COPY_GENERATION, None, "TITLE", None),
            ListingMapRow(5, FillMode.IMAGE, None, "IMAGE", None),
            ListingMapRow(6, FillMode.AI_TEXT, None, None, None),
            ListingMapRow(7, FillMode.SKIP, None, None, None),
        ]
    )
    result = overlay_columns(workbook, mapping)
    by_index = {c["column_index"]: c["config"] for c in result.columns}

    assert by_index[1]["fill_type"] == "DIRECT_MAP"
    assert by_index[1]["source"] == {"from": "SKU_MASTER", "key": "SKU"}
    assert by_index[2]["fill_type"] == "ENUM"
    assert by_index[2]["source"]["key"] == "Color"
    assert by_index[3]["fill_type"] == "ENUM"
    assert "source" not in by_index[3]
    assert by_index[4]["fill_type"] == "DIRECT_MAP"
    assert by_index[4]["source"]["attribute_name"] == "TITLE"
    assert by_index[5]["fill_type"] == "IMAGE"
    assert by_index[5]["source"]["slot"] == 1
    assert by_index[6]["fill_type"] == "AI_TEXT"
    assert by_index[7]["fill_type"] == "SKIP"
    assert "requiredness" not in by_index[1]
    assert "machine_key" not in by_index[1]


def test_overlay_repeats_generation_name_by_column_order() -> None:
    workbook = [
        _col(column_index=1, label="Highlight A"),
        _col(column_index=2, label="Highlight B"),
        _col(column_index=3, label="Bullet A"),
        _col(column_index=4, label="Bullet B"),
        _col(column_index=5, label="Image A"),
        _col(column_index=6, label="Image B"),
    ]
    mapping = _mapping(
        [
            ListingMapRow(1, FillMode.COPY_GENERATION, None, "ITEM_HIGHLIGHTS", None),
            ListingMapRow(2, FillMode.COPY_GENERATION, None, "ITEM_HIGHLIGHTS", None),
            ListingMapRow(3, FillMode.COPY_GENERATION, None, "BULLET_POINTS", None),
            ListingMapRow(4, FillMode.COPY_GENERATION, None, "BULLET_POINTS", None),
            ListingMapRow(5, FillMode.IMAGE, None, "IMAGE", None),
            ListingMapRow(6, FillMode.IMAGE, None, "IMAGE", None),
        ]
    )
    result = overlay_columns(workbook, mapping)
    by_index = {c["column_index"]: c["config"]["source"] for c in result.columns}
    assert by_index[1]["attribute_name"] == "ITEM_HIGHLIGHTS"
    assert by_index[1]["index"] == 1
    assert by_index[2]["index"] == 2
    assert by_index[3]["attribute_name"] == "BULLET_POINTS"
    assert by_index[3]["index"] == 1
    assert by_index[4]["index"] == 2
    assert by_index[5]["attribute_name"] == "IMAGE"
    assert by_index[5]["slot"] == 1
    assert by_index[6]["slot"] == 2


def test_overlay_missing_column_index_fails() -> None:
    workbook = [
        _col(column_index=1, label="Seller SKU"),
        _col(column_index=2, label="Color"),
    ]
    mapping = _mapping(
        [ListingMapRow(1, FillMode.COPY_PIM, "SKU", None, None)],
        pim_fields=[PimFieldRow("SKU", True)],
    )
    with pytest.raises(ValueError, match="no silent SKIP"):
        overlay_columns(workbook, mapping)


def test_overlay_unknown_column_index_fails() -> None:
    workbook = [_col(column_index=1, label="Seller SKU")]
    mapping = _mapping(
        [ListingMapRow(99, FillMode.COPY_PIM, "SKU", None, None)],
        pim_fields=[PimFieldRow("SKU", True)],
    )
    with pytest.raises(ValueError, match="not found in workbook"):
        overlay_columns(workbook, mapping)


def test_render_mapping_sql() -> None:
    workbook = [_col(column_index=1, label="Seller SKU")]
    mapping = _mapping(
        [ListingMapRow(1, FillMode.COPY_PIM, "SKU", None, None)],
        pim_fields=[PimFieldRow("SKU", True)],
    )
    result = overlay_columns(workbook, mapping)
    sql = render_mapping_sql(
        columns=result.columns,
        attribute_spec=result.attribute_spec,
        layout=WorkbookLayout(
            sheet_name="Template",
            header_label_row=4,
            machine_key_row=5,
            data_start_row=7,
        ),
        xlsm_name="cat.xlsm",
        mapping_name="map.xlsx",
        marketplace_id="AMAZON",
    )
    assert "UPDATE categories" in sql
    assert "--marketplace AMAZON" in sql
    assert ":category_external_id" in sql
    assert "INSERT INTO listing_template_column" in sql


@pytest.mark.skipif(not _MAP.is_file(), reason="listing mapping template xlsx missing")
def test_parse_tmp_mapping_template_examples() -> None:
    amazon = parse_mapping_workbook(_MAP, MarketplaceId.AMAZON)
    flipkart = parse_mapping_workbook(_MAP, MarketplaceId.FLIPKART)
    myntra = parse_mapping_workbook(_MAP, MarketplaceId.MYNTRA)
    assert any(r.pim_field == "SKU" for r in amazon.pim_fields)
    for mapping in (amazon, flipkart, myntra):
        assert mapping.listing_rows
        assert len(mapping.marketplace_columns) == len(mapping.listing_rows)
        modes = {row.fill_mode for row in mapping.listing_rows}
        assert FillMode.COPY_PIM in modes
        assert FillMode.IMAGE in modes
        assert FillMode.SKIP in modes


def test_parse_mapping_workbook_requires_fill_mode_for_each_marketplace_column(
    tmp_path: Path,
) -> None:
    path = tmp_path / "map.xlsx"
    _write_mapping(
        path,
        rows=[
            (1, "Seller SKU", "COPY_PIM", "SKU", "", ""),
            (2, "Color", "", "", "", ""),
        ],
    )
    with pytest.raises(ValueError, match="fill_mode required"):
        parse_mapping_workbook(path, MarketplaceId.AMAZON)


def test_parse_mapping_workbook_roundtrip(tmp_path: Path) -> None:
    path = tmp_path / "map.xlsx"
    _write_mapping(
        path,
        rows=[
            (1, "Seller SKU", "COPY_PIM", "SKU", "", ""),
            (2, "Color", "ENUM_FROM_PIM", "Color", "", ""),
        ],
    )
    mapping = parse_mapping_workbook(path, MarketplaceId.AMAZON)
    assert len(mapping.pim_fields) == 2
    assert len(mapping.marketplace_columns) == 2
    assert mapping.listing_rows[0].column_index == 1
    assert mapping.listing_rows[1].fill_mode == FillMode.ENUM_FROM_PIM


def test_build_columns_flipkart_index_and_dropdown_sheets(tmp_path: Path) -> None:
    path = tmp_path / "bedsheet.xlsx"
    wb = Workbook()
    ws_idx = wb.active
    ws_idx.title = "Index"
    ws_idx["A1"] = "Sub-categories in the file"
    ws_idx["C1"] = "Allowed Values"
    ws_idx["D1"] = "Bedsheet"
    ws_idx["A2"] = "bedsheet"
    ws_idx["D2"] = "Color"
    ws_idx["D3"] = "Red"
    ws_idx["D4"] = "Blue"
    ws = wb.create_sheet("bedsheet")
    ws["A1"] = "Seller SKU ID"
    ws["B1"] = "Country Of Origin"
    ws["C1"] = "Color"
    ws["D1"] = "Main Image URL"
    ws["A2"] = "Text"
    ws["B2"] = "Single - Text"
    ws["C2"] = "Click Here to get Allowed Values"
    ws["D2"] = "URL"
    dd = wb.create_sheet("DropDownValuesForColumn1")
    dd["A1"] = "India"
    dd["A2"] = "China"
    bogus = wb.create_sheet("DropDownValuesForColumn3")
    bogus["A1"] = "should-not-attach-to-url-column"
    wb.save(path)

    columns = build_columns(
        path,
        layout=WorkbookLayout(
            sheet_name="bedsheet",
            header_label_row=1,
            machine_key_row=2,
            data_start_row=5,
        ),
        include_requiredness=False,
    )
    by_index = {c["column_index"]: c["config"] for c in columns}
    assert by_index[1]["fill_type"] == "DIRECT_MAP"
    assert by_index[2]["fill_type"] == "ENUM"
    assert by_index[2]["valid_values"] == ["India", "China"]
    assert by_index[3]["fill_type"] == "ENUM"
    assert by_index[3]["valid_values"] == ["Red", "Blue"]
    assert by_index[4]["fill_type"] == "DIRECT_MAP"
