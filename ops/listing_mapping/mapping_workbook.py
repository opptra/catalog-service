"""Parse the restricted listing-mapping Excel workbook (tmp template contract).

Sheets:
  - pim_contract: pim_field, requirement (shared)
  - amazon_mapping / flipkart_mapping / myntra_mapping: column_index,
    marketplace_column, fill_mode on the same row. fill_mode required.

``--marketplace`` selects which mapping sheet to overlay.
"""

from __future__ import annotations

from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path

from listing_mapping.marketplace import (
    MAPPING_SHEET_NAMES,
    MarketplaceId,
    parse_marketplace_id,
)
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class FillMode(StrEnum):
    COPY_PIM = "COPY_PIM"
    COPY_GENERATION = "COPY_GENERATION"
    ENUM = "ENUM"
    # Parser aliases for older sheets; both collapse to ENUM ± pim_field.
    ENUM_FROM_PIM = "ENUM_FROM_PIM"
    ENUM_AI = "ENUM_AI"
    AI_TEXT = "AI_TEXT"
    CONSTANT = "CONSTANT"
    IMAGE = "IMAGE"
    SKIP = "SKIP"


def _validate_row_occupancy(row: ListingMapRow, *, sheet: str, excel_row: int) -> None:
    """Enforce which of pim / generation / constant may be set for each fill_mode."""
    mode = row.fill_mode
    pim = row.pim_field
    gen = row.generation
    const = row.constant_value
    loc = f"{sheet} row {excel_row} column_index={row.column_index}"

    if gen and ":" in gen:
        raise ValueError(
            f"{loc}: generation {gen!r} must be a bare name (no :n). "
            "Repeat the name on each column; numbering is by column_index order."
        )

    def _require_pim() -> None:
        if not pim:
            raise ValueError(f"{loc}: {mode.value} requires pim_field")

    def _forbid_pim() -> None:
        if pim:
            raise ValueError(f"{loc}: {mode.value} must leave pim_field blank")

    def _require_gen() -> None:
        if not gen:
            raise ValueError(f"{loc}: {mode.value} requires generation")

    def _forbid_gen() -> None:
        if gen:
            raise ValueError(f"{loc}: {mode.value} must leave generation blank")

    def _require_const() -> None:
        if not const:
            raise ValueError(f"{loc}: {mode.value} requires constant_value")

    def _forbid_const() -> None:
        if const:
            raise ValueError(f"{loc}: {mode.value} must leave constant_value blank")

    if mode == FillMode.COPY_PIM:
        _require_pim()
        _forbid_gen()
        _forbid_const()
    elif mode in {FillMode.COPY_GENERATION, FillMode.IMAGE}:
        _require_gen()
        _forbid_pim()
        _forbid_const()
    elif mode == FillMode.CONSTANT:
        _require_const()
        _forbid_pim()
        _forbid_gen()
    elif mode == FillMode.SKIP:
        _forbid_pim()
        _forbid_gen()
        _forbid_const()
    elif mode == FillMode.ENUM:
        _forbid_gen()
        _forbid_const()
        # pim_field optional
    elif mode == FillMode.ENUM_FROM_PIM:
        _require_pim()
        _forbid_gen()
        _forbid_const()
    elif mode == FillMode.ENUM_AI:
        _forbid_pim()
        _forbid_gen()
        _forbid_const()
    elif mode == FillMode.AI_TEXT:
        _forbid_gen()
        _forbid_const()
        # pim_field optional
    else:
        raise ValueError(f"{loc}: unhandled fill_mode {mode.value}")


@dataclass(frozen=True)
class PimFieldRow:
    pim_field: str
    mandatory: bool


@dataclass(frozen=True)
class MarketplaceColumnRow:
    excel_row: int
    column_index: int
    marketplace_column: str


@dataclass(frozen=True)
class ListingMapRow:
    column_index: int
    fill_mode: FillMode
    pim_field: str | None
    generation: str | None
    constant_value: str | None


@dataclass(frozen=True)
class MappingWorkbook:
    pim_fields: list[PimFieldRow]
    marketplace_columns: list[MarketplaceColumnRow]
    listing_rows: list[ListingMapRow]


_REQ_MANDATORY = frozenset({"mandatory", "required", "always", "true", "yes", "1"})
_REQ_OPTIONAL = frozenset({"optional", "false", "no", "0", ""})


def _cell_str(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith("="):
        # Unevaluated formula (e.g. status / VLOOKUP) — treat as empty for inputs.
        return ""
    return str(value).strip()


def _as_int(value: object | None) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.startswith("="):
        return None
    try:
        return int(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _header_map(ws: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        raw = ws.cell(1, col).value
        if raw is None or not str(raw).strip():
            continue
        key = " ".join(str(raw).strip().casefold().replace("_", " ").split())
        headers[key] = col
    return headers


def _require_headers(
    headers: dict[str, int],
    needed: dict[str, str],
    *,
    sheet: str,
) -> dict[str, int]:
    """needed: normalized_name → display name for errors."""
    out: dict[str, int] = {}
    for norm, display in needed.items():
        if norm not in headers:
            raise ValueError(f"Sheet {sheet!r} missing required column {display!r}")
        out[norm] = headers[norm]
    return out


def _parse_requirement(raw: str) -> bool:
    needle = raw.casefold()
    if needle in _REQ_MANDATORY:
        return True
    if needle in _REQ_OPTIONAL:
        return False
    raise ValueError(f"Invalid requirement {raw!r} (expected Mandatory/Optional)")


def _parse_pim_contract(ws: Worksheet) -> list[PimFieldRow]:
    headers = _header_map(ws)
    cols = _require_headers(
        headers,
        {"pim field": "pim_field", "requirement": "requirement"},
        sheet="pim_contract",
    )
    rows: list[PimFieldRow] = []
    seen: set[str] = set()
    for r in range(2, (ws.max_row or 1) + 1):
        field = _cell_str(ws.cell(r, cols["pim field"]).value)
        req = _cell_str(ws.cell(r, cols["requirement"]).value)
        if not field and not req:
            continue
        if not field:
            raise ValueError(f"pim_contract row {r}: pim_field is empty")
        if field in seen:
            raise ValueError(f"pim_contract row {r}: duplicate pim_field {field!r}")
        seen.add(field)
        rows.append(PimFieldRow(pim_field=field, mandatory=_parse_requirement(req)))
    if not rows:
        raise ValueError("pim_contract has no data rows")
    return rows


def _parse_marketplace_mapping_sheet(
    ws: Worksheet,
    *,
    sheet: str,
) -> tuple[list[MarketplaceColumnRow], list[ListingMapRow]]:
    headers = _header_map(ws)
    cols = _require_headers(
        headers,
        {
            "column index": "column_index",
            "marketplace column": "marketplace_column",
            "fill mode": "fill_mode",
        },
        sheet=sheet,
    )
    pim_col = headers.get("pim field")
    gen_col = headers.get("generation")
    const_col = headers.get("constant value")

    marketplace: list[MarketplaceColumnRow] = []
    listing: list[ListingMapRow] = []
    seen: set[int] = set()
    for r in range(2, (ws.max_row or 1) + 1):
        idx_raw = ws.cell(r, cols["column index"]).value
        name = _cell_str(ws.cell(r, cols["marketplace column"]).value)
        mode_raw = _cell_str(ws.cell(r, cols["fill mode"]).value)
        pim = _cell_str(ws.cell(r, pim_col).value) if pim_col else ""
        gen = _cell_str(ws.cell(r, gen_col).value) if gen_col else ""
        const = _cell_str(ws.cell(r, const_col).value) if const_col else ""
        column_index = _as_int(idx_raw)
        unused = (column_index is None or column_index == 0) and not mode_raw
        if unused and not name and not pim and not gen and not const:
            continue
        if column_index is None or column_index == 0:
            raise ValueError(f"{sheet} row {r}: column_index required")
        if column_index < 1:
            raise ValueError(f"{sheet} row {r}: column_index must be >= 1")
        if column_index in seen:
            raise ValueError(f"{sheet} row {r}: duplicate column_index {column_index}")
        seen.add(column_index)
        if not name:
            raise ValueError(
                f"{sheet} row {r}: marketplace_column is empty for column_index={column_index}"
            )
        if not mode_raw:
            raise ValueError(
                f"{sheet} row {r}: fill_mode required for column_index={column_index} ({name!r})"
            )
        try:
            fill_mode = FillMode(mode_raw)
        except ValueError as exc:
            known = ", ".join(m.value for m in FillMode)
            raise ValueError(
                f"{sheet} row {r}: invalid fill_mode {mode_raw!r}. Expected: {known}"
            ) from exc
        listing_row = ListingMapRow(
            column_index=column_index,
            fill_mode=fill_mode,
            pim_field=pim or None,
            generation=gen or None,
            constant_value=const or None,
        )
        _validate_row_occupancy(listing_row, sheet=sheet, excel_row=r)
        marketplace.append(
            MarketplaceColumnRow(
                excel_row=r,
                column_index=column_index,
                marketplace_column=name,
            )
        )
        listing.append(listing_row)
    if not marketplace:
        raise ValueError(f"{sheet} has no data rows")
    return marketplace, listing


def parse_mapping_workbook(
    path: Path,
    marketplace_id: MarketplaceId | str,
) -> MappingWorkbook:
    if not path.is_file():
        raise FileNotFoundError(f"Mapping workbook not found: {path}")
    mid = (
        marketplace_id
        if isinstance(marketplace_id, MarketplaceId)
        else parse_marketplace_id(str(marketplace_id))
    )
    sheet = MAPPING_SHEET_NAMES[mid]
    # data_only=False so openpyxl-written static cells always load (formulas skipped).
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        names = {n.casefold(): n for n in wb.sheetnames}
        if "pim_contract" not in names:
            raise ValueError("Mapping workbook missing sheet 'pim_contract'")
        if sheet.casefold() not in names:
            raise ValueError(f"Mapping workbook missing sheet {sheet!r}")
        pim = _parse_pim_contract(wb[names["pim_contract"]])
        marketplace, listing = _parse_marketplace_mapping_sheet(
            wb[names[sheet.casefold()]],
            sheet=sheet,
        )
    finally:
        wb.close()

    pim_names = {row.pim_field for row in pim}
    for row in listing:
        if row.pim_field and row.pim_field not in pim_names:
            raise ValueError(
                f"{sheet} column_index={row.column_index}: pim_field "
                f"{row.pim_field!r} is not on pim_contract"
            )
    return MappingWorkbook(
        pim_fields=pim,
        marketplace_columns=marketplace,
        listing_rows=listing,
    )


def build_attribute_spec(pim_fields: list[PimFieldRow]) -> dict[str, list[str]]:
    allowed: list[str] = []
    mandatory: list[str] = []
    seen: set[str] = set()
    for row in pim_fields:
        if row.pim_field in seen:
            continue
        seen.add(row.pim_field)
        allowed.append(row.pim_field)
        if row.mandatory:
            mandatory.append(row.pim_field)
    if "SKU" not in seen:
        allowed.insert(0, "SKU")
        mandatory.insert(0, "SKU")
    elif "SKU" not in mandatory:
        mandatory.insert(0, "SKU")
    return {"allowed": allowed, "mandatory": mandatory}
