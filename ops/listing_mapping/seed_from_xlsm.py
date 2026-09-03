"""Seed a mapping workbook from a blank marketplace .xlsm.

Fills the matching marketplace mapping sheet (``amazon_mapping`` / ``flipkart_mapping``
/ ``myntra_mapping``) from the blank workbook. Every column gets an explicit
fill_mode (no silent SKIP). Other marketplace sheets are created empty, or kept
when ``--out`` already has them.

From repo root (server venv):

  PYTHONPATH=ops:server python -m listing_mapping.seed_from_xlsm \\
    --marketplace AMAZON \\
    --xlsm tmp/listing_mapping/input/BED_LINEN_SET.xlsm \\
    --out tmp/listing_mapping/input/listing_mapping.xlsx
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

from listing_mapping import _bootstrap  # noqa: F401
from listing_mapping.build_template import MAPPING_SHEETS, build
from listing_mapping.marketplace import parse_marketplace_id
from listing_mapping.marketplace.registry import get_adapter
from openpyxl import load_workbook
from utils.listing_template_columns import build_columns


def _cell_text(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _bare_generation(value: str) -> str:
    if ":" not in value:
        return value
    return value.split(":", 1)[0].strip()


def _read_pim_rows(path: Path) -> list[tuple[str, str]]:
    if not path.is_file():
        return []
    wb = load_workbook(path, data_only=True)
    names = {str(name).casefold(): name for name in wb.sheetnames}
    if "pim_contract" not in names:
        return []
    ws = wb[names["pim_contract"]]
    rows: list[tuple[str, str]] = []
    for r in range(2, (ws.max_row or 1) + 1):
        field = _cell_text(ws.cell(r, 1).value)
        req = _cell_text(ws.cell(r, 2).value)
        if not field:
            continue
        rows.append((field, req or "Optional"))
    return rows


def _read_mapping_sheet_rows(
    path: Path, sheet_name: str
) -> list[tuple[int, str, str, str, str, str]]:
    if not path.is_file():
        return []
    wb = load_workbook(path, data_only=False)
    names = {str(name).casefold(): name for name in wb.sheetnames}
    key = sheet_name.casefold()
    if key not in names:
        return []
    ws = wb[names[key]]
    rows: list[tuple[int, str, str, str, str, str]] = []
    for r in range(2, (ws.max_row or 1) + 1):
        idx_raw = ws.cell(r, 1).value
        if idx_raw is None or idx_raw == "":
            continue
        if isinstance(idx_raw, str) and idx_raw.startswith("="):
            continue
        try:
            idx = int(idx_raw)
        except (TypeError, ValueError):
            continue
        mode = _cell_text(ws.cell(r, 3).value)
        if not mode:
            continue
        rows.append(
            (
                idx,
                _cell_text(ws.cell(r, 2).value),
                mode,
                _cell_text(ws.cell(r, 4).value),
                _bare_generation(_cell_text(ws.cell(r, 5).value)),
                _cell_text(ws.cell(r, 6).value),
            )
        )
    return rows


def _existing_marketplace_maps(
    path: Path,
) -> dict[str, list[tuple[int, str, str, str, str, str]]]:
    out: dict[str, list[tuple[int, str, str, str, str, str]]] = {}
    for marketplace_id, sheet_name in MAPPING_SHEETS:
        rows = _read_mapping_sheet_rows(path, sheet_name)
        if rows:
            out[marketplace_id] = rows
    return out


def _merge_pim_rows(
    seed_rows: list[tuple[str, str]],
    existing_rows: list[tuple[str, str]],
) -> list[tuple[str, str]]:
    if not existing_rows:
        return seed_rows
    if not seed_rows:
        return existing_rows
    seen = {name for name, _req in seed_rows}
    extra = [row for row in existing_rows if row[0] not in seen]
    return seed_rows + extra


def _is_dropdown(config: dict) -> bool:
    if config.get("fill_type") == "ENUM":
        return True
    return bool(config.get("valid_values") or config.get("valid_values_by_parent"))


def _label_key(label: str) -> str:
    return " ".join(label.casefold().split())


def _bed_linen_pim_rows() -> list[tuple[str, str]]:
    """Customer flatfile fields we expect for bed linen."""
    return [
        ("SKU", "Mandatory"),
        ("Brand", "Mandatory"),
        ("Color", "Mandatory"),
        ("Size", "Mandatory"),
        ("Material", "Optional"),
        ("Pattern", "Optional"),
        ("Thread Count", "Optional"),
        ("Number of Pieces", "Optional"),
        ("Manufacturer", "Optional"),
        ("Model Number", "Optional"),
    ]


def _bed_linen_map_rows(
    by_index: dict[int, dict],
) -> dict[int, tuple[str, str, str, str]]:
    """Explicit fill_modes for every Amazon BED_LINEN_SET column.

    Returns partial overrides; caller fills remaining indices with SKIP.
    """
    rows: dict[int, tuple[str, str, str, str]] = {}

    def add(
        idx: int,
        mode: str,
        *,
        pim: str = "",
        gen: str = "",
        const: str = "",
    ) -> None:
        if idx not in by_index:
            raise ValueError(f"Expected column_index={idx} in workbook")
        rows[idx] = (mode, pim, gen, const)

    # --- identity / variation structure ---
    add(1, "COPY_PIM", pim="SKU")
    add(2, "CONSTANT", const="BED_LINEN_SET")
    add(3, "CONSTANT", const="Edit (Partial Update)")
    add(4, "ENUM_AI")  # Parentage Level
    add(5, "SKIP")  # Parent SKU — set manually when listing children
    add(6, "ENUM_AI")  # Variation Theme Name

    # --- generated copy ---
    add(7, "COPY_GENERATION", gen="TITLE")
    add(8, "COPY_GENERATION", gen="ITEM_HIGHLIGHTS")

    # --- brand / ids ---
    add(9, "ENUM_FROM_PIM", pim="Brand")
    add(10, "SKIP")  # Product Id Type
    add(11, "SKIP")  # Product Id

    # --- browse (first node only; rest unused) ---
    add(12, "ENUM_AI")
    for idx in range(13, 17):
        add(idx, "SKIP")

    add(17, "COPY_PIM", pim="Model Number")
    add(18, "COPY_PIM", pim="Manufacturer")
    add(19, "SKIP")  # UNSPSC
    add(20, "SKIP")  # National Stock Number

    # --- images (main + 8 others; swatch unused) ---
    add(21, "IMAGE", gen="IMAGE")
    for idx in range(22, 30):
        add(idx, "IMAGE", gen="IMAGE")
    add(30, "SKIP")  # Swatch Image URL

    add(31, "COPY_GENERATION", gen="DESCRIPTION")
    for idx in range(32, 37):
        add(idx, "COPY_GENERATION", gen="BULLET_POINTS")
    add(37, "COPY_GENERATION", gen="BACKEND_KEYWORDS")

    # --- product attributes ---
    add(38, "ENUM_AI")  # Style
    add(39, "ENUM_FROM_PIM", pim="Material")
    for idx in range(40, 44):  # extra Material slots
        add(idx, "SKIP")
    add(44, "COPY_PIM", pim="Material")  # Fabric Type (free text) ← Material
    for idx in range(45, 49):
        add(idx, "SKIP")
    add(49, "COPY_PIM", pim="Number of Pieces")  # Number of Items
    add(50, "AI_TEXT")  # Item Type Name
    add(51, "ENUM_FROM_PIM", pim="Color")
    add(52, "ENUM_FROM_PIM", pim="Size")
    add(53, "COPY_PIM", pim="Number of Pieces")
    add(54, "COPY_PIM", pim="Model Number")  # Part Number
    add(55, "ENUM_AI")  # Theme (first)
    for idx in range(56, 60):
        add(idx, "SKIP")
    add(60, "ENUM_AI")  # Weave Type
    add(61, "ENUM_AI")  # Care Instructions (first)
    for idx in range(62, 66):
        add(idx, "SKIP")
    add(66, "SKIP")  # Manufacturer Contact Information

    # display dims — not in bed-linen PIM
    for idx in range(67, 78):
        add(idx, "SKIP")

    add(78, "ENUM_FROM_PIM", pim="Pattern")
    add(79, "SKIP")  # Finish Type
    add(80, "SKIP")  # Unit Count
    add(81, "SKIP")  # Unit Count Type
    add(82, "SKIP")  # Product Site Launch Date

    add(83, "ENUM_AI")  # Included Components (first)
    for idx in range(84, 88):
        add(idx, "SKIP")

    # sports / league — not bed linen
    for idx in range(88, 95):
        add(idx, "SKIP")

    add(95, "SKIP")  # External Product Information Entity
    add(96, "SKIP")
    add(97, "SKIP")  # Pillow Size
    add(98, "COPY_PIM", pim="Thread Count")
    add(99, "ENUM_AI")  # Seasons (first)
    for idx in range(100, 104):
        add(idx, "SKIP")

    # importer / packer / pillowcase / sheet geometry — skip unless later needed
    for idx in range(104, 134):
        add(idx, "SKIP")

    add(134, "SKIP")  # Is Green Purchasing Law Compliant
    add(135, "SKIP")  # Item Weight
    add(136, "SKIP")  # Item Weight Unit

    # offer / condition / gift / image locations (non-URL)
    for idx in range(137, 151):
        add(idx, "SKIP")

    # refurbished / accessories / battery — skip
    for idx in range(151, 162):
        add(idx, "SKIP")

    # inventory / pricing / B2B / shipping — ops fills outside this flow
    for idx in range(162, 195):
        add(idx, "SKIP")

    # package / item dims
    for idx in range(195, 209):
        add(idx, "SKIP")

    add(209, "ENUM_AI")  # Country of Origin

    # hazmat / GHS / age / regulation ids / compliance media
    for idx in range(210, 262):
        add(idx, "SKIP")

    # bed-linen compliance enums — model picks from constrained lists
    add(262, "ENUM_AI")  # Compliance - Bed Linen Set Components
    add(263, "ENUM_AI")  # Compliance - Is Handmade
    add(264, "ENUM_AI")  # Compliance - Printing method
    add(265, "ENUM_AI")  # Compliance Weave Type
    add(266, "ENUM_AI")  # Compliance - Outer Surface Material
    add(267, "ENUM_AI")  # Compliance - Embellishment Feature

    for idx in range(268, 273):
        add(idx, "SKIP")  # GHS Chemical H Code

    # Validate mode vs dropdown shape.
    for idx, (mode, _pim, _gen, _const) in rows.items():
        cfg = by_index[idx]
        dd = _is_dropdown(cfg)
        label = cfg.get("label")
        if mode in {"ENUM_AI", "ENUM_FROM_PIM"} and not dd:
            raise ValueError(
                f"column_index={idx} ({label!r}): {mode} requires a dropdown"
            )
        if mode in {"COPY_GENERATION", "AI_TEXT", "IMAGE"} and dd:
            raise ValueError(
                f"column_index={idx} ({label!r}): {mode} is free-text but column "
                "is a dropdown — use ENUM_AI / ENUM_FROM_PIM"
            )
    return rows


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--marketplace", required=True)
    parser.add_argument(
        "--xlsm",
        type=Path,
        required=True,
        help="Blank marketplace listing workbook (.xlsm, .xlsx, or .xls)",
    )
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument("--sheet-name", default=None)
    parser.add_argument("--header-label-row", type=int, default=None)
    parser.add_argument("--machine-key-row", type=int, default=None)
    parser.add_argument("--data-start-row", type=int, default=None)
    args = parser.parse_args(argv)

    if not args.xlsm.is_file():
        print(f"listing workbook not found: {args.xlsm}", file=sys.stderr)
        return 1

    try:
        marketplace_id = parse_marketplace_id(args.marketplace)
        adapter = get_adapter(marketplace_id)
        layout = adapter.workbook_layout(
            sheet_name=args.sheet_name,
            header_label_row=args.header_label_row,
            machine_key_row=args.machine_key_row,
            data_start_row=args.data_start_row,
        )
        columns = build_columns(args.xlsm, layout=layout, include_requiredness=False)
    except (ValueError, FileNotFoundError, OSError, NotImplementedError) as exc:
        print(f"seed_from_xlsm failed: {exc}", file=sys.stderr)
        return 1

    by_index = {int(c["column_index"]): c["config"] for c in columns}
    mkt_rows = [
        (int(c["column_index"]), str(c["config"].get("label") or "")) for c in columns
    ]

    if (
        marketplace_id.value == "AMAZON"
        and _label_key(by_index.get(2, {}).get("label", "")) == "product type"
        and _label_key(by_index.get(1, {}).get("label", "")) == "sku"
    ):
        starters = _bed_linen_map_rows(by_index)
        pim_rows = _bed_linen_pim_rows()
    else:
        print(
            "No category-specific starter map for this workbook shape; "
            "defaulting every column_index to explicit SKIP.",
            file=sys.stderr,
        )
        starters = {}
        pim_rows = [("SKU", "Mandatory")]

    # Every marketplace column gets an explicit fill_mode (SKIP unless starter).
    sheet_rows = [
        (idx, label, *(starters.get(idx, ("SKIP", "", "", ""))))
        for idx, label in mkt_rows
    ]

    maps = _existing_marketplace_maps(args.out)
    maps[marketplace_id.value] = sheet_rows
    pim_rows = _merge_pim_rows(pim_rows, _read_pim_rows(args.out))

    out = build(
        out_xlsx=args.out,
        pim_rows=pim_rows,
        marketplace_maps=maps,
    )

    counts = Counter(row[2] for row in sheet_rows)
    print(
        f"Wrote {out} ({marketplace_id.value} {len(sheet_rows)} columns; fill_modes={dict(counts)})",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
