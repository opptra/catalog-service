"""Build Excel-only listing mapping template.

Writes ``tmp/listing_mapping_template.xlsx``.

Uniqueness is by Excel ``column_index`` (not by inventing numbered labels).
``listing_map`` selects ``column_index``; ``marketplace_column`` is looked up.
"""

# ruff: noqa: E402, E501

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_REPO = Path(__file__).resolve().parents[2]
if str(_REPO / "server") not in sys.path:
    sys.path.insert(0, str(_REPO / "server"))

from entities.catalog.attribute_enums import AttributeName

OUT_XLSX = _REPO / "tmp" / "listing_mapping_template.xlsx"
OUT_DIR = _REPO / "tmp" / "listing_mapping_template"
OUT_README = _REPO / "tmp" / "listing_mapping_template_README.txt"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
ERR_FILL = PatternFill("solid", fgColor="FFC7CE")
ERR_FONT = Font(color="9C0006", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
BODY_FONT = Font(size=11)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

PIM_ROWS = 80
# Max marketplace_columns / listing_map data rows (header is row 1).
# listing_map keeps formula rows through this ceiling so a new marketplace_columns
# entry auto-appears; blank sources stay blank (no Excel 0), and fill_mode stays
# required so mapping cannot be skipped.
MKT_ROWS = 320
ARRAY_MAX = 10
SLOT_MAX = 12


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN
    ws.freeze_panes = "A2"


def _autosize(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _set(ws, row: int, col: int, value: str | int | None) -> None:
    """Write a cell without creating empty inlineStr (Excel repair trigger)."""
    cell = ws.cell(row, col)
    cell.border = THIN
    if value is None or value == "":
        return
    cell.value = value


def _list_dv(formula1: str, title: str, error: str) -> DataValidation:
    return DataValidation(
        type="list",
        formula1=formula1,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        showInputMessage=True,
        errorStyle="stop",
        errorTitle=title,
        error=error,
        promptTitle=title,
        prompt="Use the dropdown only.",
    )


def _generation_tokens() -> list[str]:
    scalar = {
        AttributeName.TITLE,
        AttributeName.DESCRIPTION,
        AttributeName.KEY_FEATURES,
        AttributeName.BACKEND_KEYWORDS,
    }
    array_indexed = {AttributeName.BULLET_POINTS, AttributeName.ITEM_HIGHLIGHTS}
    multi_slot = {AttributeName.IMAGE, AttributeName.A_PLUS}
    tokens: list[str] = []
    for name in AttributeName:
        if name in scalar:
            tokens.append(name.value)
        elif name in array_indexed:
            for i in range(1, ARRAY_MAX + 1):
                tokens.append(f"{name.value}:{i}")
            tokens.append(name.value)
        elif name in multi_slot:
            for i in range(1, SLOT_MAX + 1):
                tokens.append(f"{name.value}:{i}")
        else:
            tokens.append(name.value)
    return tokens


def _index_mirror_formula(row: int) -> str:
    """Mirror marketplace_columns.column_index onto the same listing_map row.

    Blank source cells must stay blank — a bare reference becomes 0 in Excel.
    """
    return (
        f'=IF(\'marketplace_columns\'!A{row}="","",'
        f'\'marketplace_columns\'!A{row})'
    )


def _name_lookup_formula(row: int) -> str:
    """Show marketplace_column for the mirrored column_index."""
    return (
        f'=IF(OR(A{row}="",A{row}=0),"",'
        f"IFERROR(VLOOKUP(A{row},'marketplace_columns'!$A$2:$B${MKT_ROWS},2,FALSE),\"\"))"
    )


def _status_formula(row: int, *, gen_last: int, map_last_row: int) -> str:
    """Validate listing_map row. A=column_index, C=fill_mode, D=pim, E=gen, F=constant.

    Uses nested IF only (no IFS) so Excel versions without IFS do not show #NAME?.
    Empty / unused rows (A blank or 0 and no fill_mode) stay blank.
    """
    a, c, d, e, f = f"A{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}"
    idx = f"'marketplace_columns'!$A$2:$A${MKT_ROWS}"
    fill = "'lists'!$A$2:$A$9"
    pim = f"'pim_contract'!$A$2:$A${PIM_ROWS}"
    gen = f"'lists'!$C$2:$C${gen_last}"
    # Keep the same IF nesting depth as the last known-good template.
    return (
        f'=IF(AND(OR({a}="",{a}=0),{c}=""),"",'
        f'IF(OR({a}="",{a}=0),"ERROR: column_index required",'
        f'IF(COUNTIF({idx},{a})=0,"ERROR: column_index not in marketplace_columns",'
        f'IF(COUNTIF($A$2:$A${map_last_row},{a})>1,"ERROR: duplicate column_index in listing_map",'
        f'IF({c}="","ERROR: fill_mode required",'
        f'IF(COUNTIF({fill},{c})=0,"ERROR: invalid fill_mode",'
        f'IF({c}="COPY_PIM",'
        f'IF(AND({d}<>"",COUNTIF({pim},{d})>0,{e}="",{f}=""),"OK","ERROR: COPY_PIM needs pim_field only"),'
        f'IF({c}="ENUM_FROM_PIM",'
        f'IF(AND({d}<>"",COUNTIF({pim},{d})>0,{e}="",{f}=""),"OK","ERROR: ENUM_FROM_PIM needs pim_field only"),'
        f'IF({c}="COPY_GENERATION",'
        f'IF(AND({e}<>"",COUNTIF({gen},{e})>0,{d}="",{f}=""),"OK","ERROR: COPY_GENERATION needs generation only"),'
        f'IF({c}="IMAGE",'
        f'IF(AND({e}<>"",COUNTIF({gen},{e})>0,{d}="",{f}=""),"OK","ERROR: IMAGE needs generation only"),'
        f'IF({c}="CONSTANT",'
        f'IF(AND({f}<>"",{d}="",{e}=""),"OK","ERROR: CONSTANT needs constant_value only"),'
        f'IF(OR({c}="ENUM_AI",{c}="AI_TEXT",{c}="SKIP"),'
        f'IF(AND({d}="",{e}="",{f}=""),"OK","ERROR: this mode must leave pim/generation/constant blank"),'
        f'"ERROR: unhandled fill_mode"'
        f"))))))))))))"
    )


def _readme_blocks() -> list[tuple[str, str]]:
    return [
        ("Listing mapping — how to fill this workbook", "title"),
        ("", "body"),
        ("What each sheet is for", "section"),
        (
            "1. pim_contract — fields we ask the customer (Mandatory / Optional).",
            "body",
        ),
        (
            (
                "2. marketplace_columns — one row per Excel column: column_index "
                "(unique) + marketplace_column (display name from the blank .xlsm)."
            ),
            "body",
        ),
        (
            (
                "3. listing_map — one row per marketplace_columns row (index is "
                "mirrored). Set fill_mode for every row; blank fill_mode is ERROR. "
                "Use SKIP only when you intentionally omit the column."
            ),
            "body",
        ),
        (
            "4. lists — dropdown vocabulary only. Do not edit unless you know why.",
            "body",
        ),
        ("", "body"),
        ("Fill order", "section"),
        ("A. Add customer fields on pim_contract.", "body"),
        (
            (
                "B. On marketplace_columns, enter each Template column’s Excel "
                "column_index and its header name (as shown in the blank .xlsm)."
            ),
            "body",
        ),
        (
            (
                "C. On listing_map, each marketplace_columns row already appears "
                "(column_index / marketplace_column are mirrored by formula). Set "
                "fill_mode — leave blank and status stays ERROR until you choose a "
                "mode (including SKIP). Add a new marketplace_columns row and the "
                "matching listing_map row fills in automatically; you must still "
                "set fill_mode."
            ),
            "body",
        ),
        (
            (
                "D. status must be OK (green) on every marketplace_columns row. "
                "There is no silent SKIP — every defined column needs an explicit fill_mode."
            ),
            "body",
        ),
        ("", "body"),
        ("Uniqueness", "section"),
        (
            (
                "column_index is the only unique key (Excel column number from the "
                "Template sheet). Do not invent numbered names like Other Image URL1 / 2 "
                "for uniqueness — same display name can appear on different indices."
            ),
            "body",
        ),
        ("", "body"),
        ("When to pick each fill_mode", "section"),
        (
            (
                "COPY_PIM — plain copy of a customer field (not an Amazon dropdown). "
                "Set pim_field only."
            ),
            "body",
        ),
        (
            (
                "COPY_GENERATION — copy an already-generated value "
                "(title, bullets, description, keywords). Set generation only."
            ),
            "body",
        ),
        ("IMAGE — generated image URL. Set generation to IMAGE:n or A_PLUS:n.", "body"),
        (
            (
                "AI_TEXT — free-text Amazon field (not a dropdown) with no PIM twin "
                "and not from generation."
            ),
            "body",
        ),
        ("CONSTANT — fixed string (e.g. Update). Set constant_value only.", "body"),
        ("SKIP — leave blank. Leave pim/generation/constant blank.", "body"),
        ("", "body"),
        ("Amazon dropdowns (ENUM)", "section"),
        (
            "Both ENUM modes write only values from Amazon’s allowed list. Never free text.",
            "body",
        ),
        ("", "body"),
        ("ENUM_FROM_PIM — customer has a related PIM field (set pim_field).", "body"),
        (
            (
                "  1) PIM value exact-matches an allowed option (case-insensitive) "
                "→ write that option. No model."
            ),
            "body",
        ),
        (
            (
                "  2) No match → model picks from the allowed list using the full PIM bag "
                "+ product photos (mapped field is a hint). Weak evidence → may stay blank."
            ),
            "body",
        ),
        ("", "body"),
        (
            (
                "ENUM_AI — no dedicated PIM field. Leave pim_field blank. Always model-picks "
                "from the allowed list using the full PIM bag + photos."
            ),
            "body",
        ),
        ("", "body"),
        ("When is the model involved?", "section"),
        (
            (
                "ENUM_FROM_PIM mismatch → model (list-constrained). "
                "ENUM_AI → always model (list-constrained). AI_TEXT → free text. "
                "COPY_* / IMAGE / CONSTANT / SKIP / exact ENUM match → no model."
            ),
            "body",
        ),
        ("", "body"),
        ("generation tokens", "section"),
        (
            "Bare name = whole value: TITLE, DESCRIPTION, KEY_FEATURES, BACKEND_KEYWORDS.",
            "body",
        ),
        ("NAME:n list item: BULLET_POINTS:1 …, ITEM_HIGHLIGHTS:1 …", "body"),
        ("NAME:n image slot: IMAGE:1 …, A_PLUS:1 …", "body"),
        ("", "body"),
        ("Hard rules", "section"),
        (
            "• Select column_index from the dropdown — do not type free text.",
            "body",
        ),
        (
            "• marketplace_column on listing_map is auto-filled — do not edit it.",
            "body",
        ),
        ("• Each column_index may appear at most once on listing_map.", "body"),
        ("• pim_field must exist on pim_contract when used.", "body"),
        ("• Keep pim_contract minimal.", "body"),
        ("• status = OK on every filled listing_map row before handoff.", "body"),
        ("", "body"),
        ("Open this file in Microsoft Excel Desktop (File → Open).", "body"),
    ]


def _default_pim_rows() -> list[tuple[str, str]]:
    return [
        ("SKU", "Mandatory"),
        ("Color", "Mandatory"),
        ("Size", "Mandatory"),
        ("Material", "Optional"),
        ("Brand", "Mandatory"),
        ("Pattern", "Optional"),
    ]


def _default_mkt_rows() -> list[tuple[int, str]]:
    """Illustrative only — replace from a real .xlsm before mapping."""
    return [
        (1, "SKU"),
        (2, "Product Type"),
        (3, "Listing Action"),
        (4, "Parentage Level"),
        (5, "Parent SKU"),
        (6, "Variation Theme Name"),
        (7, "Item Name"),
        (8, "Item Highlight"),
        (9, "Brand Name"),
        (21, "Main Image URL"),
        (22, "Other Image URL"),
        (31, "Product Description"),
        (32, "Bullet Point"),
        (37, "Generic Keyword"),
        (39, "Material"),
        (51, "Color"),
        (52, "Size"),
        (78, "Pattern"),
    ]


def _default_map_rows() -> list[tuple[int, str, str, str, str]]:
    """Starter rows keyed to ``_default_mkt_rows`` (Amazon-style layout)."""
    return [
        (1, "COPY_PIM", "SKU", "", ""),
        (2, "CONSTANT", "", "", "BED_LINEN_SET"),
        (3, "CONSTANT", "", "", "Edit (Partial Update)"),
        (4, "ENUM_AI", "", "", ""),
        (5, "SKIP", "", "", ""),
        (6, "ENUM_AI", "", "", ""),
        (7, "COPY_GENERATION", "", "TITLE", ""),
        (8, "COPY_GENERATION", "", "ITEM_HIGHLIGHTS:1", ""),
        (9, "ENUM_FROM_PIM", "Brand", "", ""),
        (21, "IMAGE", "", "IMAGE:1", ""),
        (22, "IMAGE", "", "IMAGE:2", ""),
        (31, "COPY_GENERATION", "", "DESCRIPTION", ""),
        (32, "COPY_GENERATION", "", "BULLET_POINTS:1", ""),
        (37, "COPY_GENERATION", "", "BACKEND_KEYWORDS", ""),
        (39, "ENUM_FROM_PIM", "Material", "", ""),
        (51, "ENUM_FROM_PIM", "Color", "", ""),
        (52, "ENUM_FROM_PIM", "Size", "", ""),
        (78, "ENUM_FROM_PIM", "Pattern", "", ""),
    ]


def build(
    *,
    out_xlsx: Path | None = None,
    out_dir: Path | None = None,
    out_readme: Path | None = None,
    pim_rows: list[tuple[str, str]] | None = None,
    mkt_rows: list[tuple[int, str]] | None = None,
    map_rows: list[tuple[int, str, str, str, str]] | None = None,
    write_sidecar_csv: bool = True,
) -> Path:
    generations = _generation_tokens()
    fill_modes = [
        "COPY_PIM",
        "COPY_GENERATION",
        "ENUM_FROM_PIM",
        "ENUM_AI",
        "AI_TEXT",
        "CONSTANT",
        "IMAGE",
        "SKIP",
    ]
    requirements = ["Mandatory", "Optional"]
    gen_last = 1 + len(generations)
    pim_rows = list(pim_rows) if pim_rows is not None else _default_pim_rows()
    mkt_rows = list(mkt_rows) if mkt_rows is not None else _default_mkt_rows()
    map_rows = list(map_rows) if map_rows is not None else _default_map_rows()
    if len(mkt_rows) > MKT_ROWS - 1:
        raise ValueError(
            f"marketplace_columns has {len(mkt_rows)} rows; raise MKT_ROWS (currently {MKT_ROWS})"
        )
    map_by_idx = {row[0]: row for row in map_rows}
    if len(map_by_idx) != len(map_rows):
        raise ValueError("listing_map map_rows contain duplicate column_index values")
    mkt_indices = [row[0] for row in mkt_rows]
    if len(set(mkt_indices)) != len(mkt_indices):
        raise ValueError("marketplace_columns contain duplicate column_index values")
    missing_map = [idx for idx in mkt_indices if idx not in map_by_idx]
    if missing_map:
        raise ValueError(
            "Every marketplace_columns column_index needs a listing_map fill_mode; "
            f"missing: {missing_map[:20]}{'…' if len(missing_map) > 20 else ''}"
        )
    extra_map = sorted(set(map_by_idx) - set(mkt_indices))
    if extra_map:
        raise ValueError(
            "listing_map has column_index values not on marketplace_columns: "
            f"{extra_map[:20]}{'…' if len(extra_map) > 20 else ''}"
        )

    out_xlsx = out_xlsx or OUT_XLSX
    out_dir = out_dir or OUT_DIR
    out_readme = out_readme or OUT_README

    wb = Workbook()

    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme.sheet_properties.tabColor = "1F4E79"
    for i, (text, kind) in enumerate(_readme_blocks(), start=1):
        cell = ws_readme.cell(i, 1)
        if text:
            cell.value = text
            cell.font = (
                TITLE_FONT if kind == "title" else SECTION_FONT if kind == "section" else BODY_FONT
            )
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws_readme.column_dimensions["A"].width = 112

    ws_lists = wb.create_sheet("lists")
    ws_lists["A1"] = "fill_mode"
    ws_lists["B1"] = "requirement"
    ws_lists["C1"] = "generation"
    for i, value in enumerate(fill_modes, start=2):
        ws_lists.cell(i, 1, value)
    for i, value in enumerate(requirements, start=2):
        ws_lists.cell(i, 2, value)
    for i, value in enumerate(generations, start=2):
        ws_lists.cell(i, 3, value)
    _style_header(ws_lists, 3)
    _autosize(ws_lists, [18, 14, 22])
    ws_lists.sheet_properties.tabColor = "808080"

    r_fill = "'lists'!$A$2:$A$9"
    r_req = "'lists'!$B$2:$B$3"
    r_gen = f"'lists'!$C$2:$C${gen_last}"
    r_pim = f"'pim_contract'!$A$2:$A${PIM_ROWS}"

    # pim_contract
    ws_pim = wb.create_sheet("pim_contract")
    pim_headers = ["pim_field", "requirement"]
    for i, header in enumerate(pim_headers, start=1):
        ws_pim.cell(1, i, header)
    for r, row in enumerate(pim_rows, start=2):
        _set(ws_pim, r, 1, row[0])
        _set(ws_pim, r, 2, row[1])
    for r in range(len(pim_rows) + 2, PIM_ROWS + 1):
        _set(ws_pim, r, 1, None)
        _set(ws_pim, r, 2, None)
    _style_header(ws_pim, 2)
    _autosize(ws_pim, [24, 14])
    dv_req = _list_dv(r_req, "requirement", "Pick Mandatory or Optional.")
    ws_pim.add_data_validation(dv_req)
    dv_req.add(f"B2:B{PIM_ROWS}")

    # marketplace_columns — unique by column_index; name is display only
    ws_mkt = wb.create_sheet("marketplace_columns")
    mkt_headers = ["column_index", "marketplace_column"]
    for i, header in enumerate(mkt_headers, start=1):
        ws_mkt.cell(1, i, header)
    for r, row in enumerate(mkt_rows, start=2):
        _set(ws_mkt, r, 1, row[0])
        _set(ws_mkt, r, 2, row[1])
    _style_header(ws_mkt, 2)
    _autosize(ws_mkt, [14, 28])

    # listing_map — formula rows through MKT_ROWS so new marketplace_columns
    # entries auto-mirror. Blank sources stay blank (no Excel 0). fill_mode is
    # blank on spare rows → status ERROR once an index appears (must map explicitly).
    # column_index / marketplace_column are formula-driven — no special fill.
    ws_map = wb.create_sheet("listing_map")
    map_headers = [
        "column_index",
        "marketplace_column",
        "fill_mode",
        "pim_field",
        "generation",
        "constant_value",
        "status",
    ]
    for i, header in enumerate(map_headers, start=1):
        ws_map.cell(1, i, header)

    map_last_row = MKT_ROWS
    for r in range(2, map_last_row + 1):
        idx_cell = ws_map.cell(r, 1)
        idx_cell.value = _index_mirror_formula(r)
        name_cell = ws_map.cell(r, 2)
        name_cell.value = _name_lookup_formula(r)
        status = ws_map.cell(r, 7)
        status.value = _status_formula(r, gen_last=gen_last, map_last_row=map_last_row)

        mkt_offset = r - 2
        if mkt_offset < len(mkt_rows):
            mapped = map_by_idx[mkt_rows[mkt_offset][0]]
            _set(ws_map, r, 3, mapped[1])
            _set(ws_map, r, 4, mapped[2])
            _set(ws_map, r, 5, mapped[3])
            _set(ws_map, r, 6, mapped[4])
            for col in (1, 2, 7):
                ws_map.cell(r, col).border = THIN

    _style_header(ws_map, 7)
    _autosize(ws_map, [14, 22, 18, 14, 22, 16, 55])

    for dv, rng in (
        (_list_dv(r_fill, "fill_mode", "Pick fill_mode only."), f"C2:C{map_last_row}"),
        (
            _list_dv(r_pim, "pim_field", "Pick from pim_contract only."),
            f"D2:D{map_last_row}",
        ),
        (
            _list_dv(r_gen, "generation", "Pick generation token only."),
            f"E2:E{map_last_row}",
        ),
    ):
        ws_map.add_data_validation(dv)
        dv.add(rng)

    ws_map.conditional_formatting.add(
        f"G2:G{map_last_row}",
        FormulaRule(formula=['LEFT(G2,5)="ERROR"'], fill=ERR_FILL, font=ERR_FONT),
    )
    ws_map.conditional_formatting.add(
        f"G2:G{map_last_row}",
        FormulaRule(formula=['G2="OK"'], fill=OK_FILL),
    )

    out_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    def write_csv(path: Path, headers: list[str], rows: list[tuple]) -> None:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerows(rows)

    if write_sidecar_csv:
        write_csv(out_dir / "pim_contract.csv", pim_headers, pim_rows)
        write_csv(out_dir / "marketplace_columns.csv", mkt_headers, mkt_rows)
        write_csv(
            out_dir / "listing_map.csv",
            ["column_index", "fill_mode", "pim_field", "generation", "constant_value"],
            map_rows,
        )
        write_csv(out_dir / "generations.csv", ["generation"], [[g] for g in generations])
        out_readme.write_text(
            "\n".join(text for text, _ in _readme_blocks()) + "\n",
            encoding="utf-8",
        )

    wb.save(out_xlsx)
    return out_xlsx


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
